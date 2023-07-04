import requests
import openpyxl
from bs4 import BeautifulSoup


class CodeTest(object):
    """
    Name: Test B:
    The fifth tab in the spreadsheet (Blog Articles) consists of a list of blog URLs. Write a program
    to extract the Article title and featured/hero image (at the right side of the title from
    webpage) and populate the title and image url in column B and C through web scraping
    process.
    """
    wb = openpyxl.load_workbook("codetest.xlsx")
    ws = wb['blog articles']

    def web_scraping(self):
        sheet = self.ws
        base_url = "https://www.google.com"

        for name in range(2, sheet.max_row + 1):
            urls = [cell.value for cell in sheet[name]]
            url = urls[0]
            reqeust = requests.get(url=url)
            soup = BeautifulSoup(reqeust.content, 'html.parser')
            # Get article title
            article = soup.find('article')
            title = article.findChild('h1')
            get_title = title.text

            # Get article feature image
            img_tag = soup.find('article')
            img = img_tag.findChild('img').attrs['src']

            # Add image to XL
            sheet.cell(row=name, column=2).value = get_title
            sheet.cell(row=name, column=3).value = img
        self.wb.save('codetest.xlsx')
        print("Title and feature images are successfully done!")  # Just for message


if __name__ == "__main__":
    obj = CodeTest()
    obj.web_scraping()

